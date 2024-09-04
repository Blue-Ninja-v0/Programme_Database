import sqlite3
import pandas as pd
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# Constants
REPORT_VERSION = "1.2.0"

# Report Configuration
REPORT_CONFIG = {
    'project_overview': {
        'include_categories': True,
        'include_location': True,
        'include_cost': True,
        'include_wbs': True,
        'include_custom_fields': True,
    },
    'task_timeline': {
        'include_constraints': True,
        'include_predecessors': True,
        'include_activity_codes': True,
        'include_resources': True,
        'include_memos': True,
        'include_custom_fields': True,
    },
    'resource_allocation': {
        'include_calendar': True,
        'include_rates': True,
        'include_categories': True,
        'include_curves': True,
        'include_custom_fields': True,
    }
}

def create_connection(db_path):
    """Create a database connection to the SQLite database specified by db_path."""
    try:
        conn = sqlite3.connect(db_path)
        return conn
    except sqlite3.Error as e:
        print(f"Error connecting to database: {e}")
        return None

def execute_query(conn, query, params=None):
    """Execute a SQL query and return the results as a DataFrame."""
    try:
        if params:
            return pd.read_sql_query(query, conn, params=params)
        else:
            return pd.read_sql_query(query, conn)
    except pd.io.sql.DatabaseError as e:
        print(f"Error executing query: {e}")
        return None

def style_worksheet(ws):
    """Apply common styles to a worksheet."""
    header_style = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_style
        cell.fill = header_fill

def adjust_column_widths(ws):
    """Adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def save_workbook(wb):
    """Save the workbook to a temporary file and return the file path."""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
        wb.save(temp_file.name)
        return temp_file.name

def generate_field_explanation(ws, row, explanations):
    """Add field explanations to the worksheet."""
    ws.cell(row=row, column=1, value="Field Explanations:")
    row += 1
    for field, explanation in explanations.items():
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=explanation)
        row += 1
    return row + 1

def generate_project_overview_report(db_path, upload_id):
    """Generate a project overview report."""
    conn = create_connection(db_path)
    if not conn:
        return None

    try:
        query = """
        SELECT 
            p.proj_id, p.proj_short_name, p.plan_start_date, p.plan_end_date, p.scd_end_date, p.proj_url
        """
        
        if REPORT_CONFIG['project_overview']['include_categories']:
            query += ", GROUP_CONCAT(DISTINCT rcv.rsrc_catg_name) AS project_categories"
        
        if REPORT_CONFIG['project_overview']['include_location']:
            query += ", p.location_id"
        
        if REPORT_CONFIG['project_overview']['include_cost']:
            query += ", SUM(tr.target_cost) AS total_target_cost"
        
        if REPORT_CONFIG['project_overview']['include_wbs']:
            query += ", GROUP_CONCAT(DISTINCT t.wbs_id) AS top_level_wbs"
        
        query += ", p.fy_start_month_num AS obs_name"
        
        if REPORT_CONFIG['project_overview']['include_custom_fields']:
            query += ", GROUP_CONCAT(DISTINCT uv.udf_type_id || ': ' || uv.udf_text) AS custom_fields"
        
        query += """
        FROM PROJECT p
        LEFT JOIN RSRCRCAT rc ON p.proj_id = rc.rsrc_id
        LEFT JOIN RCATVAL rcv ON rc.rsrc_catg_id = rcv.rsrc_catg_id AND p.xer_file_id = rcv.xer_file_id
        LEFT JOIN TASK t ON p.proj_id = t.proj_id AND p.xer_file_id = t.xer_file_id
        LEFT JOIN TASKRSRC tr ON t.task_id = tr.task_id AND p.xer_file_id = tr.xer_file_id
        LEFT JOIN UDFVALUE uv ON p.proj_id = uv.proj_id AND p.xer_file_id = uv.xer_file_id
        WHERE p.xer_file_id = ?
        GROUP BY p.proj_id
        """
        
        project_df = execute_query(conn, query, params=(upload_id,))
        if project_df is None:
            return None

        milestones_query = """
        SELECT task_name, early_start_date, early_end_date, total_float_hr_cnt
        FROM TASK 
        WHERE xer_file_id = ? AND task_type = 'Milestone'
        ORDER BY early_start_date
        """
        milestones_df = execute_query(conn, milestones_query, params=(upload_id,))
        if milestones_df is None:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Project Overview"
        
        ws.cell(row=1, column=1, value=f"Project Overview Report (v{REPORT_VERSION})")
        
        row = 3
        ws.cell(row=row, column=1, value="Project Overview")
        row += 1
        
        # Add project information
        project_info = [
            ("Project Name", 'proj_short_name'),
            ("Planned Start Date", 'plan_start_date'),
            ("Planned End Date", 'plan_end_date'),
            ("Scheduled End Date", 'scd_end_date'),
            ("Project URL", 'proj_url')
        ]
        
        for label, field in project_info:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=project_df.iloc[0][field])
            row += 1
        
        # Add conditional fields based on configuration
        conditional_fields = [
            ('include_categories', "Project Categories", 'project_categories'),
            ('include_location', "Project Location", 'location_id'),
            ('include_cost', "Total Target Cost", 'total_target_cost'),
            ('include_wbs', "Top-level WBS Elements", 'top_level_wbs')
        ]
        
        for config_key, label, field in conditional_fields:
            if REPORT_CONFIG['project_overview'][config_key]:
                ws.cell(row=row, column=1, value=label)
                value = project_df.iloc[0][field] if isinstance(field, str) else field(project_df.iloc[0])
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        ws.cell(row=row, column=1, value="OBS Name")
        ws.cell(row=row, column=2, value=project_df.iloc[0]['obs_name'])
        row += 2
        
        if REPORT_CONFIG['project_overview']['include_custom_fields']:
            ws.cell(row=row, column=1, value="Custom Fields")
            row += 1
            custom_fields = project_df.iloc[0]['custom_fields'].split(',')
            for field in custom_fields:
                ws.cell(row=row, column=1, value=field.strip())
                row += 1
            row += 1
        
        # Add milestones
        ws.cell(row=row, column=1, value="Key Milestones")
        row += 1
        milestone_headers = ["Milestone Name", "Early Start Date", "Early End Date", "Total Float (hrs)"]
        for col, header in enumerate(milestone_headers, start=1):
            ws.cell(row=row, column=col, value=header)
        row += 1
        
        for _, milestone in milestones_df.iterrows():
            for col, field in enumerate(['task_name', 'early_start_date', 'early_end_date', 'total_float_hr_cnt'], start=1):
                ws.cell(row=row, column=col, value=milestone[field])
            row += 1
        
        style_worksheet(ws)
        adjust_column_widths(ws)
        
        # Add field explanations
        explanations = {
            "Project Name": "Short name of the project",
            "Planned Start Date": "The date when the project is planned to start",
            "Planned End Date": "The date when the project is planned to end",
            "Scheduled End Date": "The current scheduled end date of the project",
            "Project URL": "Web link to the project",
            "Project Categories": "Categories assigned to the project",
            "Project Location": "Location ID of the project",
            "Total Target Cost": "Total target cost of the project",
            "Top-level WBS Elements": "Highest level Work Breakdown Structure elements",
            "OBS Name": "Fiscal year start month number (used as a proxy for OBS)",
            "Custom Fields": "User-defined fields for the project",
            "Key Milestones": "Important events or checkpoints in the project timeline"
        }
        generate_field_explanation(ws, row + 2, explanations)
        
        return save_workbook(wb)

    except Exception as e:
        print(f"An error occurred while generating the project overview report: {e}")
        return None
    finally:
        conn.close()

def generate_task_timeline_report(db_path, upload_id):
    """Generate a task timeline report."""
    conn = create_connection(db_path)
    if not conn:
        return None

    try:
        query = """
        SELECT 
            t.task_code, t.task_name, t.early_start_date, t.early_end_date, 
            t.late_start_date, t.late_end_date, t.total_float_hr_cnt, t.free_float_hr_cnt,
            t.cstr_type, t.cstr_date,
            GROUP_CONCAT(DISTINCT tp.pred_task_id || ' (' || tp.pred_type || ')') AS predecessors,
            GROUP_CONCAT(DISTINCT ac.actv_code_type_id || ': ' || ac.actv_code_name) AS activity_codes,
            COUNT(DISTINCT tr.rsrc_id) AS assigned_resources,
            tm.task_memo,
            GROUP_CONCAT(DISTINCT uv.udf_type_id || ': ' || uv.udf_text) AS custom_fields
        FROM TASK t
        LEFT JOIN TASKPRED tp ON t.task_id = tp.task_id AND t.xer_file_id = tp.xer_file_id
        LEFT JOIN TASKACTV ta ON t.task_id = ta.task_id AND t.xer_file_id = ta.xer_file_id
        LEFT JOIN ACTVCODE ac ON ta.actv_code_id = ac.actv_code_id AND t.xer_file_id = ac.xer_file_id
        LEFT JOIN TASKRSRC tr ON t.task_id = tr.task_id AND t.xer_file_id = tr.xer_file_id
        LEFT JOIN TASKMEMO tm ON t.task_id = tm.task_id AND t.xer_file_id = tm.xer_file_id
        LEFT JOIN UDFVALUE uv ON t.task_id = uv.fk_id AND t.xer_file_id = uv.xer_file_id
        WHERE t.xer_file_id = ?
        GROUP BY t.task_id
        ORDER BY t.early_start_date
        """
        
        tasks_df = execute_query(conn, query, params=(upload_id,))
        if tasks_df is None:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Task Timeline"
        
        headers = ["Task Code", "Task Name", "Early Start Date", "Early End Date", "Late Start Date", "Late End Date",
                   "Total Float (hrs)", "Free Float (hrs)", "Constraint Type", "Constraint Date", "Predecessors",
                   "Activity Codes", "Assigned Resources", "Task Memo", "Custom Fields"]
        ws.append(headers)
        
        for _, row in tasks_df.iterrows():
            ws.append([row[header.lower().replace(' ', '_').replace('(', '').replace(')', '')] for header in headers])

        style_worksheet(ws)
        adjust_column_widths(ws)

        # Add legend for constraint types and predecessor types
        ws.append([])
        ws.append(["Constraint Types:"])
        ws.append(["CS - Start On", "CF - Finish On", "MSO - Start On or After", "MFO - Finish On or After"])
        ws.append(["SNET - Start No Earlier Than", "FNET - Finish No Earlier Than", "SNLT - Start No Later Than", "FNLT - Finish No Later Than"])
        
        ws.append([])
        ws.append(["Predecessor Types:"])
        ws.append(["FS - Finish to Start", "SS - Start to Start", "FF - Finish to Finish", "SF - Start to Finish"])

        return save_workbook(wb)

    except Exception as e:
        print(f"An error occurred while generating the task timeline report: {e}")
        return None
    finally:
        conn.close()

def generate_resource_allocation_report(db_path, upload_id):
    """Generate a resource allocation report for the given project."""
    conn = create_connection(db_path)
    if not conn:
        return None

    try:
        query = """
        SELECT 
            r.rsrc_id, r.rsrc_name, t.task_name, tr.target_qty, tr.act_reg_qty, tr.remain_qty,
            c.clndr_name, rr.cost_per_qty, 
            GROUP_CONCAT(DISTINCT rcv.rsrc_catg_name) AS resource_categories,
            tr.curv_id AS resource_curve,
            GROUP_CONCAT(DISTINCT uv.udf_type_id || ': ' || uv.udf_text) AS custom_fields
        FROM TASKRSRC tr
        JOIN RSRC r ON tr.rsrc_id = r.rsrc_id AND tr.xer_file_id = r.xer_file_id
        JOIN TASK t ON tr.task_id = t.task_id AND tr.xer_file_id = t.xer_file_id
        LEFT JOIN CALENDAR c ON r.clndr_id = c.clndr_id AND r.xer_file_id = c.xer_file_id
        LEFT JOIN RSRCRATE rr ON r.rsrc_id = rr.rsrc_id AND r.xer_file_id = rr.xer_file_id
        LEFT JOIN RSRCRCAT rc ON r.rsrc_id = rc.rsrc_id AND r.xer_file_id = rc.xer_file_id
        LEFT JOIN RCATVAL rcv ON rc.rsrc_catg_id = rcv.rsrc_catg_id AND r.xer_file_id = rcv.xer_file_id
        LEFT JOIN UDFVALUE uv ON r.rsrc_id = uv.fk_id AND r.xer_file_id = uv.xer_file_id
        WHERE tr.xer_file_id = ?
        GROUP BY tr.taskrsrc_id
        ORDER BY r.rsrc_name, t.task_name
        """
        
        resource_allocation_df = execute_query(conn, query, params=(upload_id,))
        if resource_allocation_df is None:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = "Resource Allocation"
        
        headers = ["Resource Name", "Task Name", "Target Quantity", "Actual Regular Quantity", "Remaining Quantity",
                   "Calendar", "Cost per Quantity", "Resource Categories", "Resource Curve", "Custom Fields"]
        ws.append(headers)
        
        for _, row in resource_allocation_df.iterrows():
            ws.append([row[header.lower().replace(' ', '_')] for header in headers])

        style_worksheet(ws)
        adjust_column_widths(ws)

        return save_workbook(wb)

    except Exception as e:
        print(f"An error occurred while generating the resource allocation report: {e}")
        return None
    finally:
        conn.close()

def generate_combined_report(db_path, upload_id):
    """Generate a combined report including project overview, task timeline, and resource allocation."""
    try:
        wb = Workbook()
        
        # Generate Project Overview
        overview_path = generate_project_overview_report(db_path, upload_id)
        if overview_path:
            overview_wb = load_workbook(overview_path)
            overview_ws = overview_wb.active
            wb.create_sheet("Project Overview")
            for row in overview_ws.iter_rows(values_only=True):
                wb["Project Overview"].append(row)
        
        # Generate Task Timeline
        timeline_path = generate_task_timeline_report(db_path, upload_id)
        if timeline_path:
            timeline_wb = load_workbook(timeline_path)
            timeline_ws = timeline_wb.active
            wb.create_sheet("Task Timeline")
            for row in timeline_ws.iter_rows(values_only=True):
                wb["Task Timeline"].append(row)
        
        # Generate Resource Allocation
        resource_path = generate_resource_allocation_report(db_path, upload_id)
        if resource_path:
            resource_wb = load_workbook(resource_path)
            resource_ws = resource_wb.active
            wb.create_sheet("Resource Allocation")
            for row in resource_ws.iter_rows(values_only=True):
                wb["Resource Allocation"].append(row)
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        return save_workbook(wb)

    except Exception as e:
        print(f"An error occurred while generating the combined report: {e}")
        return None

if __name__ == "__main__":
    import sys
    if len(sys.argv) != 3:
        print("Usage: python reports.py <database_path> <upload_id>")
        sys.exit(1)
    
    db_path = sys.argv[1]
    upload_id = sys.argv[2]
    
    combined_report_path = generate_combined_report(db_path, upload_id)
    if combined_report_path:
        print(f"Combined report generated successfully: {combined_report_path}")
    else:
        print("Failed to generate combined report")